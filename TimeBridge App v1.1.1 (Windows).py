"""
Created on Thu Aug 21 08:13:42 2025

@author: davidmenezes
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procore Timecard Downloader + Excel pasting
- Gets access token from refresh token
- Downloads timecard entries for a date range
- Removes "Cost Code Long Name" (if present), keeps columns A-K
- Writes rows (row 2..end, cols A..K) for each employee into the sheet named for that employee
  in an existing Excel workbook (preserves cell formatting by only writing values).
"""

import os
import json
import re
import io
import csv
import webbrowser
import requests
import pandas as pd
import time
import tempfile
from openpyxl import load_workbook
from datetime import datetime
from urllib.parse import urlparse, parse_qs
import traceback
from pathlib import Path
from http.server import BaseHTTPRequestHandler, HTTPServer
import threading
import io, csv, base64, secrets, getpass
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.backends import default_backend
import tkinter.simpledialog as simpledialog
from PIL import Image, ImageTk
from tkinter import simpledialog



# -------------------- CONFIG --------------------
CLIENT_ID     = 'yHTFNlKjEMAjFcnEbGNIXRN13Bq6ezt6YJVM1YLeQyY'
CLIENT_SECRET = 'bsKFfy4jbFkq9avooml9uFq2tlL97VfMGFhaaFej9l8'
# IMPORTANT: This must EXACTLY match one of the Redirect URIs registered for your app
REFRESH_TOKEN = None   # e.g. '5dR...'; set to None to force interactive (or let saved token be used)
# in the Procore Developer Portal (My Apps -> Edit App -> Redirect URIs).
# It can be a dummy like "https://example.com/callback" — you will copy the code from the browser.
REDIRECT_URI  = "http://localhost:8080/callback"
COMPANY_ID    = '9993'
TOKEN_FILE    = "procore_token.json"
DOWNLOADS = Path.home() / "Downloads"
DOWNLOADS.mkdir(parents=True, exist_ok=True)
OUT_CSV = DOWNLOADS / "timecards.csv"
CREDS_FILE = Path.home() / ".procore_timecard_users.bin"   # encrypted user store
KDF_ITERATIONS = 200_000                                   # PBKDF2 rounds

# ------------------------------------------------


# ---------------- Login Prompt ------------------
def build_login_screen(root, on_success):
    root.title("TimeBridge - Secure User Login")
    root.geometry("400x500")
    root.configure(bg="#f0f0f0")

    # --- Container Frame ---
    frame = tk.Frame(root, bg="#ffffff", bd=0)
    frame.place(relx=0.5, rely=0.5, anchor="center", width=340, height=500)

    # --- Company Logo (from URL with local fallback), preserve aspect ratio ---
    logo_url = "https://tse2.mm.bing.net/th/id/OIP.GEYLp9at-_1NnvGGsOMUjgHaDU?rs=1&pid=ImgDetMain&o=7&rm=3"
    max_size = (110, 110)  # maximum footprint for logo (width, height)

    def _make_logo_photo(image_bytes):
        img = Image.open(io.BytesIO(image_bytes)).convert("RGBA")
        img.thumbnail(max_size, Image.LANCZOS)  # preserves aspect ratio
        # create a square/transparent background and paste the resized logo centered
        bg = Image.new("RGBA", max_size, (255, 255, 255, 0))
        x = (max_size[0] - img.width) // 2
        y = (max_size[1] - img.height) // 2
        bg.paste(img, (x, y), img)
        return ImageTk.PhotoImage(bg)

    logo_img = None
    try:
        r = requests.get(logo_url, timeout=8)
        r.raise_for_status()
        logo_img = _make_logo_photo(r.content)
    except Exception:
        # fallback to local file if remote load fails
        logo_path = "company_logo.png"
        if os.path.exists(logo_path):
            try:
                with open(logo_path, "rb") as f:
                    logo_img = _make_logo_photo(f.read())
            except Exception:
                logo_img = None

    if logo_img:
        logo_label = tk.Label(frame, image=logo_img, bg="#ffffff")
        logo_label.image = logo_img   # keep reference
        logo_label.pack(pady=(18, 10))


   # --- App title / company / developer lines (below logo, above username) ---
    tk.Label(frame, text="TimeBridge", font=("Helvetica", 18, "bold"),
         bg="#ffffff", fg="#000000").pack(pady=(4,0))
    tk.Label(frame, text="Zlotnick Construction, Inc. (ZCI)", font=("Helvetica", 11),
         bg="#ffffff", fg="#333333").pack()
    tk.Label(frame, text="Developed by David M. Menezes", font=("Helvetica", 9, "italic"),
         bg="#ffffff", fg="#666666").pack(pady=(0,10))

    # --- Username Entry ---
    tk.Label(frame, text="Username", font=("Helvetica", 10), bg="#ffffff", fg="black").pack(pady=(30, 5))
    username_entry = tk.Entry(frame, width=30, font=("Helvetica", 11))
    username_entry.pack(ipady=5, pady=(0, 15))   # space below username field

    tk.Label(frame, text="Password", font=("Helvetica", 10),
         bg="#ffffff", fg="black").pack(pady=(10, 5))
    password_entry = tk.Entry(frame, show="*", width=30, font=("Helvetica", 11))
    password_entry.pack(ipady=5, pady=(0, 20))  

    # --- define attempt_login inside build_login_screen ---
    def attempt_login():
        user = username_entry.get().strip()
        pwd = password_entry.get().strip()

        if validate_user(user, pwd):   # your AES-256 check
            root.destroy()
            on_success()
        else:
            tk.messagebox.showerror("Login Failed", "Invalid username or password")

    # --- Login Button ---
    login_btn = tk.Button(
    frame,
    text="Login",
    width=30,
    font=("Helvetica", 11, "bold"),
    bg="#bbbbbb",
    fg="black",
    command=attempt_login
    )
    login_btn.pack(pady=(30, 12), ipady=8)   # <-- ipady increases button height


    return username_entry, password_entry
# ------------------------------------------------


# ------------------ Utilities -------------------
def log_error_to_file(error: Exception):
    """
    Saves error details to a timestamped log file in ~/Documents/ErrorLogs/
    """
    docs = Path.home() / "Documents" / "ErrorLogs"
    docs.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_file = docs / f"error_{timestamp}.txt"

    with open(log_file, "w", encoding="utf-8") as f:
        f.write("=== Procore Timecard Downloader Error ===\n")
        f.write(f"Timestamp: {timestamp}\n\n")
        f.write("Error Message:\n")
        f.write(str(error) + "\n\n")
        f.write("Traceback:\n")
        f.write(traceback.format_exc())

    return log_file

def save_tokens(tokens):
    with open(TOKEN_FILE, "w") as f:
        json.dump(tokens, f)
    print(f"Saved tokens to {TOKEN_FILE}")

def load_tokens():
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "r") as f:
            return json.load(f)
    return None

def sanitize_sheet_name(name: str) -> str:
    if not isinstance(name, str):
        name = str(name)
    name = re.sub(r'[:\\/?\*\[\]]+', ' ', name).strip()
    return name[:31] if len(name) > 31 else (name or "Unknown")

def append_log(log_box, text: str):
    """Append text to log_box from any thread safely."""
    def _append():
        log_box.insert(tk.END, text + "\n")
        log_box.see(tk.END)
    try:
        log_box.after(0, _append)
    except Exception:
        # fallback if log_box not ready — ignore
        pass
# ------------------------------------------------

# --------------- OAuth helpers ------------------
def exchange_refresh_for_access(refresh_token):
    url = "https://login.procore.com/oauth/token"
    data = {
        "grant_type": "refresh_token",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "refresh_token": refresh_token
    }
    return requests.post(url, data=data)

def exchange_code_for_tokens(code):
    url = "https://login.procore.com/oauth/token"
    data = {
        "grant_type": "authorization_code",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "code": code,
        "redirect_uri": REDIRECT_URI
    }
    r = requests.post(url, data=data)
    r.raise_for_status()
    return r.json()

def interactive_get_new_tokens():
    class OAuthHandler(BaseHTTPRequestHandler):
        def do_GET(self):
            parsed = urlparse(self.path)
            qs = parse_qs(parsed.query)
            self.server.auth_code = qs.get("code", [None])[0]

            self.send_response(200)
            self.send_header('Content-type', 'text/html')
            self.end_headers()
            self.wfile.write(b"<html><body><h2>Authorization successful. You can close this window.</h2></body></html>")

        def log_message(self, format, *args):
            return  # suppress logging to stdout

    auth_url = (
        "https://login.procore.com/oauth/authorize"
        f"?response_type=code&client_id={CLIENT_ID}&redirect_uri={REDIRECT_URI}"
    )
    webbrowser.open(auth_url)

    # Start local server in another thread
    server = HTTPServer(('localhost', 8080), OAuthHandler)
    thread = threading.Thread(target=server.handle_request)
    thread.start()
    thread.join()

    code = server.auth_code
    if not code:
        raise SystemExit("No authorization code received. Try again.")

    tokens = exchange_code_for_tokens(code)
    save_tokens(tokens)
    return tokens

def get_access_token():
    # priority: REFRESH_TOKEN var -> saved token file -> interactive flow
    if REFRESH_TOKEN:
        print("Using REFRESH_TOKEN from script variable to get access token...")
        r = exchange_refresh_for_access(REFRESH_TOKEN)
        if r.status_code == 200:
            tokens = r.json()
            save_tokens(tokens)
            return tokens["access_token"]
        else:
            print("Refresh-token-in-script failed:", r.status_code, r.text)

    tokens = load_tokens()
    if tokens and "refresh_token" in tokens:
        print("Attempting to refresh access token using saved refresh_token...")
        r = exchange_refresh_for_access(tokens["refresh_token"])
        if r.status_code == 200:
            new_tokens = r.json()
            save_tokens(new_tokens)
            return new_tokens["access_token"]
        else:
            print("Saved refresh failed:", r.status_code, r.text)

    print("No valid refresh token available — starting manual auth...")
    tokens = interactive_get_new_tokens()
    return tokens["access_token"]
# ------------------------------------------------

# ---------------- API calls (company-level) -----------
def try_timecards_endpoint(access_token, start_date, end_date):
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }
    url = f"https://api.procore.com/rest/v1.0/companies/{COMPANY_ID}/timecard_entries"
    params = {
        "company_id": COMPANY_ID,
        "start_date": start_date,
        "end_date": end_date,
        "include_totals": "true"
    }
    r = requests.get(url, headers=headers, params=params)
    print(f"API Response: {r.status_code} - {r.text}")  # More logging
    if r.status_code == 200:
        try:
            data = r.json()
            if not data.get('timecards'):
                print("No timecards found for this date range.")
                return None
            df = pd.json_normalize(data['timecards'])
            return df
        except Exception as e:
            log_error_to_file(e)
            return None
    else:
        print(f"/timecards returned {r.status_code}: {r.text}")
        log_error_to_file(f"API error: {r.text}")
        return None



def try_timecard_entries_endpoint(access_token, start_date, end_date):
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }
    url = f"https://api.procore.com/rest/v1.0/companies/{COMPANY_ID}/timecard_entries"
    params = {
        "company_id": COMPANY_ID,
        "start_date": start_date,
        "end_date": end_date,
        "include_totals": "true"
    }
    r = requests.get(url, headers=headers, params=params)
    print(f"API Response: {r.status_code} - {r.text}")  # Add more details here
    if r.status_code == 200:
        try:
            data = r.json()
            df = pd.json_normalize(data)
            print(f"Timecard Data: {data}")  # Inspect the data returned
            return df
        except Exception as e:
            print(f"Error parsing JSON: {str(e)}")
            with open(OUT_CSV_RAW, "wb") as f:
                f.write(r.content)
            try:
                return pd.read_csv(OUT_CSV_RAW)
            except Exception as e:
                print(f"Error reading CSV: {str(e)}")
                return None
    else:
        print(f"/timecards returned {r.status_code}: {r.text}")
        return None


def try_reports_export(access_token, start_date, end_date):
    """
    Export the 'Timecard Report' from Procore Reports API as CSV,
    normalize headers/encoding, drop Cost Code Long Name (col G), keep A..K,
    then save to OUT_CSV and return the cleaned DataFrame.
    """
    headers = {"Authorization": f"Bearer {access_token}", "Accept": "application/json"}
    list_url = "https://api.procore.com/rest/v1.0/reports"
    r = requests.get(list_url, headers=headers, params={"company_id": COMPANY_ID})
    if r.status_code != 200:
        print(f"/reports list returned {r.status_code}: {r.text}")
        return None

    reports = r.json()
    candidates = [rep for rep in reports if "timecard" in (rep.get("name","") or "").lower()]
    if not candidates:
        print("No 'Timecard Report' found.")
        return None

    report_id = candidates[0].get("id") or candidates[0].get("report_id")
    export_url = f"https://api.procore.com/rest/v1.0/reports/{report_id}/export"
    params = {
        "company_id": COMPANY_ID,
        "start_date": start_date,
        "end_date": end_date,
        "format": "csv"
    }

    r2 = requests.get(export_url, headers={"Authorization": f"Bearer {access_token}"}, params=params, timeout=60)
    if r2.status_code != 200:
        print(f"Report export failed {r2.status_code}: {r2.text}")
        return None

    content = r2.content

    # 1) Try pandas with utf-8-sig
    df = None
    try:
        df = pd.read_csv(io.BytesIO(content), encoding="utf-8-sig")
    except Exception:
        # 2) Fallback: detect delimiter and try again with text decode
        text = None
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = content.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            print("Unable to decode CSV from report export.")
            return None
        sample = "\n".join(text.splitlines()[:10])
        try:
            delim = csv.Sniffer().sniff(sample, delimiters=[",",";","\t","|"]).delimiter
        except Exception:
            delim = ","
        try:
            df = pd.read_csv(io.StringIO(text), sep=delim, engine="python")
        except Exception as e:
            print("Failed to parse exported CSV:", e)
            return None

    # Normalize column names: strip whitespace
    df.columns = [str(c).strip() for c in df.columns]

    # Drop columns matching Cost Code Long Name (case-insensitive)
    cols_to_drop = [c for c in df.columns if "cost" in c.lower() and "long" in c.lower() and "name" in c.lower()]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop, errors='ignore')
    else:
        # defensive: if 7th column looks like cost code long name, drop it
        if df.shape[1] >= 7:
             c7 = df.columns[6]
             if "cost" in str(c7).lower() and "long" in str(c7).lower() and "name" in str(c7).lower():
                df = df.drop(columns=[c7], errors='ignore')


    # Truncate to A..K (first 11 columns) if more exist
    if df.shape[1] > 11:
        df = df.iloc[:, :11]

    # Save final CSV with BOM so Excel opens cleanly
    try:
        df.to_csv(str(OUT_CSV), index=False, encoding="utf-8-sig")
    except Exception as e:
        print("Failed writing cleaned CSV:", e)
        return None

    print(f"Saved cleaned report to {OUT_CSV}")
    return df


# ----------------------------------------------------

# --------------- Clean and paste -------------------
def clean_timecard_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    cols_to_drop = [c for c in df.columns if c.lower().strip() in (
        "cost code long name", "cost_code_long_name", "cost_code_longname")]
    for c in cols_to_drop:
        df = df.drop(columns=[c], errors='ignore')
    if df.shape[1] > 11:
        df = df.iloc[:, :11]
    employee_cols = [c for c in df.columns if c.lower().strip() in ("employee name","employee_name","employee")]
    if not employee_cols:
        fallback = [c for c in df.columns if 'employee' in c.lower()]
        if fallback:
            employee_cols = [fallback[0]]
    if employee_cols:
        df = df.rename(columns={employee_cols[0]: "Employee Name"})
    else:
        df["Employee Name"] = "Unknown"
    return df

def paste_into_workbook(df: pd.DataFrame, workbook_path: str):
    if df is None or df.empty:
        print("No data to paste.")
        return
    if not os.path.exists(workbook_path):
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    wb = load_workbook(workbook_path)
    grouped = df.groupby("Employee Name")
    for employee, group in grouped:
        sheet_name = sanitize_sheet_name(employee)
        if sheet_name not in wb.sheetnames:
            print(f"Creating sheet: {sheet_name}")
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        if ws.max_row >= 2:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
        rows = group.values.tolist()
        start_row = 2
        for r_idx, row_vals in enumerate(rows, start=start_row):
            for c_idx, value in enumerate(row_vals, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(workbook_path)
    print(f"Pasted timecards into workbook: {workbook_path}")
# ----------------------------------------------------

# -------------------- Main -------------------------
def main():
    print("1) Obtain access token (will refresh or run manual auth if needed)...")
    access_token = get_access_token()

    start_date = input("Enter pay period START date (YYYY-MM-DD, Monday): ").strip()
    end_date   = input("Enter pay period END date (YYYY-MM-DD, Sunday): ").strip()
    try:
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")
    except Exception:
        print("Invalid date format. Use YYYY-MM-DD.")
        return

    print("2) Trying company-level /timecards endpoint...")
    df = try_timecards_endpoint(access_token, start_date, end_date)

    if df is None or df.empty:
        print("3) Falling back to /timecard_entries endpoint...")
        df = try_timecard_entries_endpoint(access_token, start_date, end_date)

    if df is None or df.empty:
        print("4) Falling back to Reports API export (searching for a 'Timecard' report)...")
        df = try_reports_export(access_token, start_date, end_date)

    if df is None or df.empty:
        print("No timecard data returned from API. If your account requires the UI 'Timecard Report' export,")
        print("please export the CSV manually and put it at:", OUT_CSV_RAW)
        return


    excel_path = input("Enter path to Excel workbook to paste into (e.g. timecards.xlsx): ").strip()
    if not excel_path:
        print("No workbook provided; exiting.")
        return

    try:
        paste_into_workbook(df_clean, excel_path)
    except Exception as e:
        print("Failed to paste into workbook:", e)
        print("You can manually open:", OUT_CSV_CLEAN)
        return

    print("Done.")



import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
from datetime import datetime
import threading

# import your existing functions (from your code above)
# e.g. get_access_token, try_timecards_endpoint, try_timecard_entries_endpoint, try_reports_export,
# clean_timecard_df, paste_into_workbook, etc.


# --------- AES-256 Encryption -------------
def _pbkdf2_key(passphrase: str, salt: bytes, length: int = 32) -> bytes:
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(), length=length, salt=salt,
        iterations=KDF_ITERATIONS, backend=default_backend()
    )
    return kdf.derive(passphrase.encode("utf-8"))

def _encrypt_json(obj: dict, passphrase: str) -> bytes:
    salt = secrets.token_bytes(16)
    key = _pbkdf2_key(passphrase, salt, 32)  # AES-256
    aes = AESGCM(key)
    nonce = secrets.token_bytes(12)
    plaintext = json.dumps(obj).encode("utf-8")
    ct = aes.encrypt(nonce, plaintext, None)
    # pack: salt | nonce | ciphertext
    return b"v1" + salt + nonce + ct

def _decrypt_json(blob: bytes, passphrase: str) -> dict:
    if not blob.startswith(b"v1"):
        raise ValueError("Unsupported credentials format")
    blob = blob[2:]
    salt, nonce, ct = blob[:16], blob[16:28], blob[28:]
    key = _pbkdf2_key(passphrase, salt, 32)
    aes = AESGCM(key)
    plaintext = aes.decrypt(nonce, ct, None)
    return json.loads(plaintext.decode("utf-8"))

def _new_user_record(username: str, password: str) -> dict:
    salt = secrets.token_bytes(16)
    pwd_hash = _pbkdf2_key(password, salt, 32)
    return {
        "u": username,
        "algo": "pbkdf2-sha256",
        "iter": KDF_ITERATIONS,
        "salt": base64.b64encode(salt).decode("ascii"),
        "hash": base64.b64encode(pwd_hash).decode("ascii"),
    }

def _verify_user(password: str, rec: dict) -> bool:
    if rec.get("algo") != "pbkdf2-sha256":
        return False
    iters = int(rec.get("iter", KDF_ITERATIONS))
    salt = base64.b64decode(rec["salt"])
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(), length=32, salt=salt,
        iterations=iters, backend=default_backend()
    )
    try:
        kdf.verify(password.encode("utf-8"), base64.b64decode(rec["hash"]))
        return True
    except Exception:
        return False
    
def reset_user_store():
    try:
        # Prompt for master key before allowing reset
        key = simpledialog.askstring(
            "Master Key Required",
            "Enter the master key to reset your username & password:",
            show="*"
        )
        if key is None:  # user canceled
            return

        master_key_file = Path.home() / ".procore_master.key"
        if not master_key_file.exists():
            messagebox.showerror("Error", "No stored master key found. Cannot verify.")
            return

        stored_master = master_key_file.read_text(encoding="utf-8")
        if key != stored_master:
            messagebox.showerror("Error", "Invalid master key. Reset aborted.")
            return

        if messagebox.askyesno("Confirm Reset", 
                               "This will DELETE all saved users and Master Key. Continue?"):
            user_store = Path.home() / ".procore_timecard_users.bin"
            master_key = Path.home() / ".procore_master.key"

            if user_store.exists():
                user_store.unlink()
            if master_key.exists():
                master_key.unlink()

            messagebox.showinfo("Reset Complete", 
                                "User store and Master Key reset.\nRestart the app to reinitialize.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to reset username and password:\n{e}")



# ------------------------------------------


# ------ Secure Store Bootstrap/Load -------
def _bootstrap_creds(master_key: str, root_win=None):
    # Ask for initial admin user/password
    username = simpledialog.askstring("Setup", "Create admin username:", parent=root_win)
    if not username:
        raise SystemExit("Setup cancelled")
    password = simpledialog.askstring("Setup", "Create admin password:", show="*", parent=root_win)
    if not password:
        raise SystemExit("Setup cancelled")
    store = {"users": [_new_user_record(username.strip(), password)]}
    # write encrypted user store (uses AESGCM + PBKDF2 as implemented)
    CREDS_FILE.write_bytes(_encrypt_json(store, master_key))
    # also persist the master key to a file so we can decrypt later (we read it back in validate_user)
    master_key_file = Path.home() / ".procore_master.key"
    master_key_file.write_text(master_key, encoding="utf-8")


def _load_creds(master_key: str) -> dict:
    blob = CREDS_FILE.read_bytes()
    return _decrypt_json(blob, master_key)
def validate_user(username: str, password: str) -> bool:
    """
    Validate username/password against the encrypted user store using
    _load_creds (AESGCM decryption) and _verify_user (PBKDF2 password check).
    """
    try:
        if not CREDS_FILE.exists():
            return False
        master_key_file = Path.home() / ".procore_master.key"
        if not master_key_file.exists():
            return False
        master_key = master_key_file.read_text(encoding="utf-8")
        store = _load_creds(master_key)   # returns {"users":[...]}
        rec = next((u for u in store.get("users", []) if u.get("u") == username), None)
        if not rec:
            return False
        return _verify_user(password, rec)
    except Exception:
        # don't leak details to UI — just return False
        return False



# ------------------------------------------





# ------------- GUI Helpers ----------------
def require_login(root) -> bool:
    # Ask for master key to decrypt store (admin-defined secret)
    master_key = simpledialog.askstring("Security", "Enter Master Key:", show="*", parent=root)
    if not master_key:
        messagebox.showerror("Auth", "No master key provided.")
        return False

    if not CREDS_FILE.exists():
        if messagebox.askyesno("Setup", "No user store found. Initialize now?"):
            _bootstrap_creds(master_key, root)
        else:
            return False

    try:
        store = _load_creds(master_key)
    except Exception:
        messagebox.showerror("Auth", "Invalid master key or corrupt store.")
        return False

    # User login prompt
    username = simpledialog.askstring("Login", "Username:", parent=root)
    password = simpledialog.askstring("Login", "Password:", show="*", parent=root)
    if not username or not password:
        return False

    # find user
    rec = next((u for u in store.get("users", []) if u.get("u") == username), None)
    if not rec or not _verify_user(password, rec):
        messagebox.showerror("Auth", "Invalid username or password.")
        return False

    # optional: allow adding users if admin logged in
    # (uncomment if desired)
    # if messagebox.askyesno("Users", "Add another user?"):
    #     new_u = simpledialog.askstring("Users", "New username:", parent=root)
    #     new_p = simpledialog.askstring("Users", "New password:", show="*", parent=root)
    #     if new_u and new_p:
    #         store["users"].append(_new_user_record(new_u, new_p))
    #         CREDS_FILE.write_bytes(_encrypt_json(store, master_key))

    return True

def run_timecard_process(start_date, end_date, excel_path, log_box, push_btn=None):
    try:
        # disable push button while download runs (if provided)
        if push_btn:
            push_btn.config(state="disabled")

        append_log(log_box, "1) Getting access token...")
        access_token = get_access_token()

        # validate dates
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
        except Exception:
            messagebox.showerror("Date Error", "Dates must be in YYYY-MM-DD format")
            return

        append_log(log_box, "2) Trying company-level /timecards endpoint...")
        df = try_timecards_endpoint(access_token, start_date, end_date)

        if df is None or df.empty:
            append_log(log_box, "Falling back to /timecard_entries...")
            df = try_timecard_entries_endpoint(access_token, start_date, end_date)

        if df is None or df.empty:
            append_log(log_box, "Falling back to Reports API export...")
            df = try_reports_export(access_token, start_date, end_date)

        if df is None or df.empty:
            messagebox.showwarning("No Data", "No timecard data returned from API.\nPlease export CSV manually.")
            return

        # Clean and save CSV
        df_clean = clean_timecard_df(df)
        df_clean.to_csv(str(OUT_CSV), index=False, encoding="utf-8-sig")
        append_log(log_box, f"Saved cleaned report to {OUT_CSV}")

        # enable push button now that CSV exists
        if push_btn:
            push_btn.config(state="normal")
            append_log(log_box, "Push to Viewpoint button enabled")

        # Optionally paste into Excel (existing flow) -- unchanged, but use df_clean
        if excel_path:
            append_log(log_box, f"Pasting into workbook: {excel_path}")
            paste_into_workbook(df_clean, excel_path)
            append_log(log_box, f"Pasted data into {excel_path}")
            messagebox.showinfo("Success", "Timecards successfully processed!")
        else:
            append_log(log_box, "No workbook provided; skipping paste.")
            messagebox.showinfo("Success", f"CSV exported to: {OUT_CSV}")

    except Exception as e:
        log_path = log_error_to_file(e)
        messagebox.showerror("Error", f"Process failed:\n{e}\n\nError saved to:\n{log_path}")
        # ensure push button is left disabled in error case
        if push_btn:
            push_btn.config(state="disabled")


def start_process(start_entry, end_entry, excel_var, log_box, push_btn=None):
    start_date = start_entry.get().strip()
    end_date = end_entry.get().strip()
    excel_path = excel_var.get().strip()

    # run in thread so UI doesn’t freeze
    threading.Thread(
        target=run_timecard_process,
        args=(start_date, end_date, excel_path, log_box, push_btn),
        daemon=True
    ).start()


def select_excel(excel_var):
    path = filedialog.askopenfilename(
        title="Select Excel or CSV File", 
        filetypes=[("All Files", "*.*"),  # Allows all file types
                  ("CSV files", "*.csv"),  # Allow CSV files
                  ("Excel files", "*.xlsx *.xls *.xlsm *.xltx")]  # Allow all Excel formats
    )
    if path:
        excel_var.set(path)

def push_to_viewpoint_ui(csv_path, log_box, btn):
    """
    Read csv_path, then simulate typing each row's columns into the active Viewpoint window.
    User must focus/click the first input cell in Viewpoint before pressing OK in the prompt.
    """
    try:
        # disable button while running
        btn.config(state="disabled")
        append_log(log_box, "Preparing to push to Viewpoint...")

        if not Path(csv_path).exists():
            append_log(log_box, f"No CSV found at {csv_path}")
            btn.config(state="normal")
            return

        # Prompt user to ensure Viewpoint is ready and focused on the first cell
        ok = messagebox.askokcancel(
            "Ready Viewpoint",
            "Please open Viewpoint, navigate to the PR Timecard Entry screen, create a new batch and\n"
            "click the FIRST cell where the first time entry should be entered. Do NOT use the keyboard/mouse during automation.\n\n"
            "Click OK when ready to begin."
        )
        if not ok:
            append_log(log_box, "Viewpoint push cancelled by user.")
            btn.config(state="normal")
            return

        # Small delay to give the user time to focus the window
        append_log(log_box, "Starting automation in 3 seconds. Do not touch keyboard/mouse...")
        time.sleep(3)

        # load csv
        df = pd.read_csv(csv_path, dtype=str).fillna("")
        total = len(df.index)
        append_log(log_box, f"Rows to push: {total}")

        # Slight pause between key strokes
        pyautogui.PAUSE = 0.08
        pyautogui.FAILSAFE = True  # move mouse to corner to abort

        for idx, row in enumerate(df.itertuples(index=False), start=1):
            # For each column in order, type value then Tab.
            # IMPORTANT: The Viewpoint input order must match the CSV column order (A..K).
            for val in row:
                s = str(val) if val is not None else ""
                if s != "":
                    pyautogui.typewrite(s)
                # tab to next field
                pyautogui.press('tab')
                time.sleep(0.02)
            # after finishing columns for this entry, press Enter to commit (or adjust if your flow uses another key)
            pyautogui.press('enter')
            append_log(log_box, f"Pushed row {idx}/{total}")
            # small pause between rows
            time.sleep(0.05)

        append_log(log_box, "Finished pushing all rows to Viewpoint.")
        messagebox.showinfo("Done", "Finished pushing to Viewpoint. Verify entries in Viewpoint.")

    except Exception as e:
        append_log(log_box, f"Error during Viewpoint push: {e}")
        messagebox.showerror("Error", f"Viewpoint automation failed:\n{e}")
    finally:
        # re-enable the button
        try:
            btn.config(state="normal")
        except Exception:
            pass

# ------------- Main UI --------------------
def launch_ui():
    root = tk.Tk()
    root.title("TimeBridge")

    tk.Label(root, text="Start Date (YYYY-MM-DD):").grid(row=0, column=0, sticky="w", padx=5, pady=5)
    start_entry = tk.Entry(root, width=15)
    start_entry.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(root, text="End Date (YYYY-MM-DD):").grid(row=1, column=0, sticky="w", padx=5, pady=5)
    end_entry = tk.Entry(root, width=15)
    end_entry.grid(row=1, column=1, padx=5, pady=5)

    excel_var = tk.StringVar()
    tk.Label(root, text="Excel Workbook:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
    tk.Entry(root, textvariable=excel_var, width=40).grid(row=2, column=1, padx=5, pady=5)
    tk.Button(root, text="Browse...", command=lambda: select_excel(excel_var)).grid(row=2, column=2, padx=5, pady=5)

    log_box = scrolledtext.ScrolledText(root, width=60, height=15, state="normal")
    log_box.grid(row=3, column=0, columnspan=3, padx=5, pady=10)

    # Push to Viewpoint button (left of Download button); starts disabled until download succeeds
    push_btn = tk.Button(
        root,
        text="Push to Viewpoint",
        state="disabled",
        command=lambda: threading.Thread(
            target=push_to_viewpoint_ui,
            args=(str(OUT_CSV), log_box, push_btn),
            daemon=True
        ).start()
    )
    push_btn.grid(row=4, column=0, padx=5, pady=10)

    run_btn = tk.Button(
        root,
        text="Download Timesheet from Procore",
        command=lambda: start_process(start_entry, end_entry, excel_var, log_box, push_btn)
    )
    run_btn.grid(row=4, column=1, pady=10)

    reset_btn = tk.Button(root, text="Reset Username & Password", command=reset_user_store, fg="red")
    reset_btn.grid(row=4, column=2, pady=10)


    root.mainloop()

if __name__ == "__main__":
    # Ensure tkinter dialogs can be shown for setup
    root_tmp = tk.Tk()
    root_tmp.withdraw()

    # First-run: if there is no CREDS_FILE, prompt for a master key and bootstrap
    if not CREDS_FILE.exists():
        master_key = simpledialog.askstring("Setup", "Create a Master Key (keep this safe):", show="*", parent=root_tmp)
        if not master_key:
            messagebox.showerror("Setup", "Master Key is required for initial setup.")
            root_tmp.destroy()
            raise SystemExit("Master Key required")
        _bootstrap_creds(master_key, root_tmp)

    root_tmp.destroy()

    # Launch sleek login screen (this uses validate_user())
    def start_main_ui():
        launch_ui()

    root = tk.Tk()

    # Download logo from URL
    #url = "https://media.glassdoor.com/sql/2611777/zlotnick-construction-squarelogo-1563561276394.png"
    #response = requests.get(url)
    #img_data = io.BytesIO(response.content)

    # Load into PIL
    #img = Image.open(img_data)

    # --- Set window/taskbar icon (Tkinter requires .ico) ---
    #with tempfile.NamedTemporaryFile(delete=False, suffix=".ico") as tmp_ico:
    #    img.save(tmp_ico, format="ICO")
    #    root.iconbitmap(tmp_ico.name)

    # --- Set login screen logo (PhotoImage) ---
    #logo_img = ImageTk.PhotoImage(img.resize((150, 150)))  # resize if needed
    #logo_label = tk.Label(root, image=logo_img, bg="white")
    #logo_label.image = logo_img
    #logo_label.pack(pady=10)

    build_login_screen(root, start_main_ui)
    root.mainloop()
